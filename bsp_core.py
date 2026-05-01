"""bsp_core - núcleo de cálculo do BSP.

A ideia foi tirar tudo o que é matemática, parsing e estatística para um sítio
sem tkinter/matplotlib/reportlab pelo meio. Assim posso correr os testes sem
abrir janela e o backend web (FastAPI) importa daqui directamente.

    python -m bsp_core --testes
    from bsp_core import calcular, PROTO_ARCO, comparar_grupos
"""

from __future__ import annotations

import os
import re
import math
import json
import logging
import unicodedata
import datetime
from typing import Optional, List, Dict, Tuple, Any, Callable

import numpy as np
from scipy import stats as _sp_stats

_logger = logging.getLogger('bsp_core')

__all__ = [
    # Constantes
    'VERSAO', 'PROG',
    'PROTO_FMS', 'PROTO_UNIPODAL', 'PROTO_TIRO', 'PROTO_ARCO',
    'PROTOCOLOS', 'TAREFAS_FUNC', 'AOM_CONFIG',
    'G_ACEL',
    # Math
    'calcular', 'assimetria', 'flagrar_outliers',
    # Parsing
    'ler_ficheiro', '_ler_ficheiro_arco', '_detectar_formato_arco',
    # Loaders
    'carregar_atletas_ref', 'atletas_ref_por_id',
    'carregar_confirmacao_arco',
    # Filename matching
    'achar_ficheiro_arco',
    # Demographics
    'comparar_grupos', 'correlacao_demografica',
    'percentis_subgrupo', 'correlacao_score',
    # Testes
    'run_testes',
]

# ═════════════════════════════════════════════════════════════════════
# Constantes
# ═════════════════════════════════════════════════════════════════════

VERSAO = "1.0"
PROG   = "Biomechanical Stability Program"

G_ACEL = 9.80665  # m/s^2

# Protocolos
PROTO_FMS      = "fms"
PROTO_UNIPODAL = "unipodal"
PROTO_TIRO     = "tiro"        # Tiro (espingarda, ISCPSI)
PROTO_ARCO     = "tiro_arco"   # Tiro com Arco (janela unica, demografia)

# Configuracao centralizada (equivalente a AOM_CONFIG em estabilidade_gui)
AOM_CONFIG = {
    "n_ens_default":       5,
    "outlier_z_thresh":    4.5,
    "outlier_pct_thresh":  0.50,
    "min_ensaios_validos": 3,
    "vel_metodo":          "classico",   # "classico" | "filtro" | "spline"
    "vel_fc_hz":           10.0,
    "vel_fs_hz":           25.0,
    "vel_spline_s":        0.0,
    "tiro_selection_ativo": True,
    "fft_ativo":            False,
    "peso_norm":            False,
}

PROTOCOLOS = {
    PROTO_FMS: {
        'nome':        'FMS Bipodal',
        'descr':       '5 ensaios por pe, bipodal',
        'lados':       [('dir', 'dir_', 0), ('esq', 'esq_', 5)],
        'n_ens':       5,
        'assimetria':  True,
        'two_windows': False,
    },
    PROTO_UNIPODAL: {
        'nome':        'Apoio Unipodal',
        'descr':       '5 ensaios por pe, unipodal',
        'lados':       [('dir', 'dir_', 0), ('esq', 'esq_', 5)],
        'n_ens':       5,
        'assimetria':  False,
        'two_windows': False,
    },
    PROTO_TIRO: {
        'nome':        'Tiro',
        'descr':       '5 ensaios bipodal, 2 janelas por ensaio',
        'lados':       [('pos', 'tiro_', 0), ('disp', 'tiro_', 0)],
        'n_ens':       5,
        'assimetria':  False,
        'two_windows': True,
    },
    PROTO_ARCO: {
        'nome':        'Tiro com Arco',
        'descr':       'Ate 30 ensaios bipodal, janela unica',
        'lados':       [('arco', 'arco_', 0)],
        'n_ens':       30,
        'assimetria':  False,
        'two_windows': False,
    },
}

TAREFAS_FUNC = {
    PROTO_TIRO: PROTOCOLOS[PROTO_TIRO],
    PROTO_ARCO: PROTOCOLOS[PROTO_ARCO],
}

# Padroes de ficheiros a ignorar (formato nao suportado pelo platform export)
_SKIP_PATTERNS = ['entire plate roll off', 'roll off']

# Chi2(0.95, df=2) - Schubert & Kirchner (2013)
_CHI2_95 = _sp_stats.chi2.ppf(0.95, df=2)

# Pontos pre-calculados da elipse (120 pontos + fecho)
_ELL_COS = np.cos(np.linspace(0.0, 2.0 * math.pi, 121))
_ELL_SIN = np.sin(np.linspace(0.0, 2.0 * math.pi, 121))


# ═════════════════════════════════════════════════════════════════════
# Deteccao e parsing de ficheiros
# ═════════════════════════════════════════════════════════════════════

def _is_skip_file(caminho: str) -> bool:
    """True se o ficheiro deve ser ignorado (roll-off exports)."""
    nome_lower = os.path.basename(caminho).lower()
    return any(pat in nome_lower for pat in _SKIP_PATTERNS)


def _detectar_formato_arco(caminho: str) -> bool:
    """
    Sniff: True se o ficheiro tem cabecalho do novo formato Tiro com Arco
    ('Entire plate COF X' na primeira ~40 linhas).
    """
    try:
        with open(caminho, 'r', encoding='iso-8859-1') as f:
            head = ''.join(next(f, '') for _ in range(40))
    except Exception:
        return False
    head_low = head.lower()
    return ('entire plate cof x' in head_low
            or 'stability export for measurement' in head_low)


def _ler_ficheiro_arco(caminho: str) -> Dict:
    """
    Parser do novo formato Tiro com Arco (tab-separated .xls text, 50 Hz).

    Cabecalho tipico:
      Stability export for measurement: ...
      Patient name: ...
      Measurement done on 03-06-2025
      Frame\tTime (ms)\tEntire plate COF X\t...

    Devolve dict com chaves: paciente, medicao, data, inicio_ms, fim_ms,
    frames (lista de {frame, t_ms, x, y, forca, dist, sel_*}).
    """
    with open(caminho, 'r', encoding='iso-8859-1') as f:
        linhas = f.readlines()

    info = {'paciente': None, 'medicao': None, 'data': None,
            'inicio_ms': None, 'fim_ms': None, 'frames': []}

    idx_dados = None
    col_idx: Dict[str, int] = {}
    for i, linha in enumerate(linhas):
        s = linha.strip()
        if s.startswith('Stability export for measurement:'):
            info['medicao'] = s[len('Stability export for measurement:'):].strip()
        elif s.lower().startswith('patient name:'):
            info['paciente'] = s.split(':', 1)[1].strip()
        elif s.startswith('Measurement done on'):
            info['data'] = s[len('Measurement done on'):].strip()
        elif s.startswith('Frame\t') or s.lower().startswith('frame\ttime'):
            headers = [h.strip().lower() for h in s.split('\t')]
            for ci, h in enumerate(headers):
                # Indexar pelo cabecalho completo
                col_idx[h] = ci
                # ... e tambem pela versao sem sufixo de unidade (ex: '(mm)', '(n)')
                #     - 'entire plate cof x (mm)' -> 'entire plate cof x'
                #     - 'time (ms)' -> 'time'
                base = re.sub(r'\s*\([^)]*\)\s*$', '', h).strip()
                if base and base != h and base not in col_idx:
                    col_idx[base] = ci
            idx_dados = i + 1
            break

    if idx_dados is None:
        raise ValueError(f"Formato arco inesperado (cabecalho ausente): "
                         f"{os.path.basename(caminho)}")

    def _find_col(*nomes) -> Optional[int]:
        """Match exact, depois prefix, depois substring (lowercase)."""
        # 1. Exact match
        for n in nomes:
            if n in col_idx:
                return col_idx[n]
        # 2. Prefix match (cobre ' ml/s', ' (m^2)' etc nao previstos acima)
        for n in nomes:
            for h, ci in col_idx.items():
                if h.startswith(n):
                    return ci
        # 3. Substring (ultimo recurso)
        for n in nomes:
            for h, ci in col_idx.items():
                if n in h:
                    return ci
        return None

    c_frame = _find_col('frame')
    c_time  = _find_col('time (ms)', 'time', 't_ms', 'time_ms')
    c_x     = _find_col('entire plate cof x', 'cof x', 'x')
    c_y     = _find_col('entire plate cof y', 'cof y', 'y')
    c_for   = _find_col('entire plate cof force', 'cof force', 'force')
    c_lx = _find_col('left selection cof x', 'left sel x', 'left cof x')
    c_ly = _find_col('left selection cof y', 'left sel y', 'left cof y')
    c_rx = _find_col('right selection cof x', 'right sel x', 'right cof x')
    c_ry = _find_col('right selection cof y', 'right sel y', 'right cof y')

    if c_x is None or c_y is None:
        raise ValueError(f"Formato arco sem colunas CoP X/Y: "
                         f"{os.path.basename(caminho)}")

    def _sf(cols, i):
        if i is None or i >= len(cols):
            return None
        try:
            v = cols[i].strip()
            return float(v) if v else None
        except (TypeError, ValueError):
            return None

    frames = []
    for linha in linhas[idx_dados:]:
        s = linha.rstrip('\n')
        if not s.strip():
            continue
        cols = s.split('\t')
        try:
            f_data = {
                'frame': (int(float(cols[c_frame]))
                          if c_frame is not None and c_frame < len(cols)
                          else len(frames)),
                't_ms': (float(cols[c_time])
                         if c_time is not None and c_time < len(cols)
                         else len(frames) * 20.0),
                'x':    float(cols[c_x]),
                'y':    float(cols[c_y]),
            }
        except (TypeError, ValueError, IndexError):
            continue
        f_data['forca']     = _sf(cols, c_for)
        f_data['dist']      = None
        f_data['sel_esq_x'] = _sf(cols, c_lx)
        f_data['sel_esq_y'] = _sf(cols, c_ly)
        f_data['sel_dir_x'] = _sf(cols, c_rx)
        f_data['sel_dir_y'] = _sf(cols, c_ry)
        frames.append(f_data)

    if frames:
        info['inicio_ms'] = int(frames[0]['t_ms'])
        info['fim_ms']    = int(frames[-1]['t_ms'])

    info['frames'] = frames
    n_sel = sum(1 for f in frames
                if f.get('sel_dir_x') is not None and f.get('sel_dir_y') is not None)
    info['tem_selection'] = (n_sel >= len(frames) * 0.5) if frames else False
    info['_formato_arco'] = True
    return info


def ler_ficheiro(caminho: str) -> Dict:
    """
    Leitor com sniffing. Se o cabecalho indicar novo formato arco,
    delega a _ler_ficheiro_arco; caso contrario nao suporta (bsp_core
    so implementa o parser arco; o legacy fica em estabilidade_gui.py).
    """
    if _is_skip_file(caminho):
        raise ValueError(f"Ficheiro ignorado (roll-off): "
                         f"{os.path.basename(caminho)}")
    if _detectar_formato_arco(caminho):
        return _ler_ficheiro_arco(caminho)
    raise ValueError("bsp_core so suporta o formato Tiro com Arco. "
                     "Para .xlsx ou formato legacy use estabilidade_gui.ler_ficheiro.")


# ═════════════════════════════════════════════════════════════════════
# Matcher de ficheiros
# ═════════════════════════════════════════════════════════════════════

def achar_ficheiro_arco(pasta: str, pid: str, trial: int) -> Optional[str]:
    """
    Procura ficheiro Stability export de arco pelo padrao:
      {pid}_{trial} - DD-MM-YYYY - Stability export.xls

    Regex case-insensitive, tolerante a espacos e a '_export' vs ' export'.
    """
    if not os.path.isdir(pasta):
        return None
    rgx = re.compile(
        rf'^{re.escape(str(pid))}_0*{int(trial)}\s*-\s*'
        rf'\d{{2}}-\d{{2}}-\d{{4}}\s*-\s*'
        rf'Stability[\s_]+export\.xls$',
        re.IGNORECASE,
    )
    for nome in os.listdir(pasta):
        if nome.startswith('._'):   # metadados macOS
            continue
        if rgx.match(nome):
            return os.path.join(pasta, nome)
    # Fallback: match relaxado (so pid_trial + stability export)
    rgx_fallback = re.compile(
        rf'^{re.escape(str(pid))}_0*{int(trial)}\b.*stability.*export',
        re.IGNORECASE,
    )
    for nome in os.listdir(pasta):
        if nome.startswith('._'):   # metadados macOS
            continue
        if rgx_fallback.match(nome):
            return os.path.join(pasta, nome)
    return None


# ═════════════════════════════════════════════════════════════════════
# Velocidades e calculo principal
# ═════════════════════════════════════════════════════════════════════

def _calcular_velocidades_eixo(x: np.ndarray, y: np.ndarray,
                                t_s: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    """
    Diferencas centradas de 2a ordem (Prieto et al., 1996):
      d/dt f(ti) = [f(ti+1) - f(ti-1)] / (ti+1 - ti-1)

    Metodo "filtro"/"spline" disponiveis via AOM_CONFIG['vel_metodo'].
    """
    metodo = AOM_CONFIG.get('vel_metodo', 'classico')
    n = len(x)

    if metodo == 'filtro' and n >= 4:
        try:
            from scipy.signal import butter, filtfilt
            fs_est = (n - 1) / (t_s[-1] - t_s[0]) if t_s[-1] > t_s[0] else AOM_CONFIG.get('vel_fs_hz', 25.0)
            nyq = fs_est / 2.0
            fc  = min(AOM_CONFIG.get('vel_fc_hz', 10.0), nyq * 0.9)
            b, a = butter(4, fc / nyq, btype='low', analog=False)
            xf = filtfilt(b, a, x)
            yf = filtfilt(b, a, y)
            dt_c = t_s[2:] - t_s[:-2]
            dt_c = np.where(dt_c == 0, 0.04, dt_c)
            return (xf[2:] - xf[:-2]) / dt_c, (yf[2:] - yf[:-2]) / dt_c
        except Exception:
            pass

    if metodo == 'spline' and n >= 4:
        try:
            from scipy.interpolate import UnivariateSpline
            _, uniq = np.unique(t_s, return_index=True)
            t_u = t_s[uniq]; x_u = x[uniq]; y_u = y[uniq]
            k = min(3, len(t_u) - 1)
            sx = UnivariateSpline(t_u, x_u, k=k, s=AOM_CONFIG.get('vel_spline_s', 0.0), ext=3)
            sy = UnivariateSpline(t_u, y_u, k=k, s=AOM_CONFIG.get('vel_spline_s', 0.0), ext=3)
            return sx.derivative()(t_s), sy.derivative()(t_s)
        except Exception:
            pass

    # Classico (default)
    dt_c = t_s[2:] - t_s[:-2]
    dt_c = np.where(dt_c == 0, 0.04, dt_c)
    return (x[2:] - x[:-2]) / dt_c, (y[2:] - y[:-2]) / dt_c


def calcular(frames, t_ini=None, t_fim=None, peso_kg=None, altura_m=None):
    """
    Metricas de estabilidade para um ensaio.

    Args:
      frames   : lista de dicts com 't_ms', 'x', 'y'
      t_ini/t_fim : janela em ms (None = sinal completo)
      peso_kg  : massa corporal (reservado para compatibilidade)
      altura_m : altura (reservado para compatibilidade)

    Returns: dict com todas as metricas ou None se len(sel) < 5.

    Ref.: Schubert & Kirchner (2013), Prieto et al. (1996), Winter (1995),
          Quijoux et al. (2021), Maurer & Peterka (2005).
    """
    sel = ([f for f in frames if t_ini <= f['t_ms'] <= t_fim]
           if t_ini is not None else list(frames))
    if len(sel) < 5:
        return None

    x = np.array([f['x'] for f in sel])
    y = np.array([f['y'] for f in sel])
    t = np.array([f['t_ms'] for f in sel])

    amp_x   = float(x.max() - x.min())
    amp_y   = float(y.max() - y.min())
    t_total = (t[-1] - t[0]) / 1000.0

    passos  = np.sqrt(np.diff(x)**2 + np.diff(y)**2)
    desl    = float(passos.sum())
    vel_med = desl / t_total if t_total > 0 else 0.0

    t_s = t / 1000.0
    vx_i, vy_i = _calcular_velocidades_eixo(x, y, t_s)
    vel_x = float(np.abs(vx_i).mean())
    vel_y = float(np.abs(vy_i).mean())
    vel_pico_x = float(np.abs(vx_i).max()) if len(vx_i) > 0 else 0.0
    vel_pico_y = float(np.abs(vy_i).max()) if len(vy_i) > 0 else 0.0

    # Elipse 95% via eigenvalores
    cov  = np.cov(x, y)
    eigv = np.linalg.eigvalsh(cov)
    ev1  = max(float(eigv[1]), 0.0)
    ev2  = max(float(eigv[0]), 0.0)
    sa   = math.sqrt(ev1) * math.sqrt(_CHI2_95)
    sb   = math.sqrt(ev2) * math.sqrt(_CHI2_95)
    area = math.pi * (sa / 2) * (sb / 2)
    ang  = math.atan2(float(eigv[1]) - float(cov[0, 0]), float(cov[0, 1]))
    mx   = float(x.mean()); my = float(y.mean())
    ca   = math.cos(ang);   sa2 = math.sin(ang)
    xe_arr = sa * _ELL_COS
    ye_arr = sb * _ELL_SIN
    ell_x = (ca * xe_arr - sa2 * ye_arr + mx).tolist()
    ell_y = (sa2 * xe_arr + ca * ye_arr + my).tolist()

    # RMS (Quijoux et al., 2021)
    x_c = x - mx; y_c = y - my
    rms_x = float(np.sqrt(np.mean(x_c**2)))
    rms_y = float(np.sqrt(np.mean(y_c**2)))
    rms_r = float(np.sqrt(np.mean(x_c**2 + y_c**2)))

    # Derivadas
    ratio_ml_ap = amp_x / amp_y if amp_y > 0 else None
    ratio_vel   = vel_x / vel_y if vel_y > 0 else None
    stiff_x = vel_med / amp_x if amp_x > 0 else None
    stiff_y = vel_med / amp_y if amp_y > 0 else None

    # Normalizacoes por dimensoes corporais (autor-definidas; seguem a logica do
    # pendulo invertido de Winter, 1995: remover efeito do tamanho corporal para
    # comparacoes inter-sujeito). Requerem peso_kg / altura_m.
    #   ea95_norm    : area elipse / altura (mm2/m)  — escala linear por altura
    #   amp_norm_x/y : amplitude / altura (mm/m)
    #   vel_norm     : velocidade media / altura (mm/s/m)
    #   stiff_mass_x/y: stiffness / massa (1/(s·kg)) — remove efeito da massa
    #   stiff_norm_x/y: stiffness / altura (1/(s·m)) — remove efeito da altura
    ea95_norm    = area / altura_m          if altura_m else None
    amp_norm_x   = amp_x / altura_m        if altura_m else None
    amp_norm_y   = amp_y / altura_m        if altura_m else None
    vel_norm     = vel_med / altura_m      if altura_m else None
    stiff_mass_x = stiff_x / peso_kg       if (stiff_x is not None and peso_kg) else None
    stiff_mass_y = stiff_y / peso_kg       if (stiff_y is not None and peso_kg) else None
    stiff_norm_x = stiff_x / altura_m      if (stiff_x is not None and altura_m) else None
    stiff_norm_y = stiff_y / altura_m      if (stiff_y is not None and altura_m) else None

    return {
        'amp_x': amp_x, 'amp_y': amp_y,
        'vel_x': vel_x, 'vel_y': vel_y, 'vel_med': vel_med,
        'vel_pico_x': vel_pico_x, 'vel_pico_y': vel_pico_y,
        'desl': desl, 'time': t_total,
        'ea95': area, 'leng_a': sa, 'leng_b': sb,
        'rms_x': rms_x, 'rms_y': rms_y, 'rms_r': rms_r,
        'ratio_ml_ap': ratio_ml_ap, 'ratio_vel': ratio_vel,
        'stiff_x': stiff_x, 'stiff_y': stiff_y,
        'ea95_norm':    ea95_norm,
        'amp_norm_x':   amp_norm_x,   'amp_norm_y':   amp_norm_y,
        'vel_norm':     vel_norm,
        'stiff_mass_x': stiff_mass_x, 'stiff_mass_y': stiff_mass_y,
        'stiff_norm_x': stiff_norm_x, 'stiff_norm_y': stiff_norm_y,
        'cof_x': x.tolist(), 'cof_y': y.tolist(),
        'mean_x': mx, 'mean_y': my,
        'ell_x': ell_x, 'ell_y': ell_y,
        't_ms': t.tolist(),
    }


def assimetria(vd: Optional[float], ve: Optional[float]) -> Optional[float]:
    """
    Indice de assimetria Dir/Esq (ratio de simetria standard em biomecânica):
      AI = (VD - VE) / [(VD + VE)/2] * 100 [%]
    """
    if vd is None or ve is None:
        return None
    med = (vd + ve) / 2.0
    return round((vd - ve) / med * 100.0, 2) if med != 0 else 0.0


def flagrar_outliers(mets_lista, chave='ea95',
                     z_thresh: Optional[float] = None,
                     pct_thresh: Optional[float] = None) -> List[bool]:
    """
    Modified z-score baseado em MAD (Iglewicz & Hoaglin, 1993):
      MZ = 0.6745 * |x - mediana| / MAD
    Flagra se MZ > z_thresh E |x - mediana|/mediana > pct_thresh.
    """
    if z_thresh   is None: z_thresh   = AOM_CONFIG.get('outlier_z_thresh',   4.5)
    if pct_thresh is None: pct_thresh = AOM_CONFIG.get('outlier_pct_thresh', 0.50)
    vals = [m[chave] if (m is not None and chave in m) else None for m in mets_lista]
    nums = np.array([v for v in vals if v is not None])
    if len(nums) < 3:
        return [False] * len(vals)
    med = np.median(nums)
    mad = np.median(np.abs(nums - med))
    if mad == 0 or med == 0:
        return [False] * len(vals)
    out = []
    for v in vals:
        if v is None:
            out.append(False)
        else:
            z   = abs(0.6745 * (v - med) / mad)
            pct = abs(v - med) / med
            out.append(bool(z > z_thresh and pct > pct_thresh))
    return out


# ═════════════════════════════════════════════════════════════════════
# Loaders (Excel)
# ═════════════════════════════════════════════════════════════════════

def _normalizar_cabecalho(s: str) -> str:
    """Normaliza cabecalho: lowercase, sem acentos, sem espacos extra."""
    if not isinstance(s, str):
        return ''
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.lower().strip()


def _normalizar_nome_sheet(nome: str) -> str:
    """Normaliza nome de sheet tolerando cp1252-mangled names."""
    if not isinstance(nome, str):
        return ''
    # Remove caracteres de replacement cp1252 ('?' na zona accentuada)
    s = nome.replace('\ufffd', '')
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.lower().strip()


def carregar_atletas_ref(caminho):
    """Lê o ficheiro de referência demográfica do tiro com arco.

    O ficheiro tem 1 linha por atleta. As colunas estão numa ordem fixa
    (ID, PESO, ALTURA, IDADE, ESTILO, CATEGORIA, GENERO, depois P1..P30,
    P_TOTAL, d1..d30). O resto das colunas tem valores manuais que não
    interessam aqui.

    Não há coluna "Nome" no ficheiro - o ID identifica e o nome vem
    da pasta (ex: '101_Tiago_Matos').
    """
    from openpyxl import load_workbook  # só importamos quando preciso

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active

    # Helpers locais para coerção de valores. Como há sempre células
    # vazias e estilos de digitação variados, vale a pena ser tolerante.
    def num(v):
        if v in (None, ''):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def int_or(v):
        n = num(v)
        return int(n) if n is not None else None

    def code_to_genero(v):
        # 1 = masculino, 2 = feminino. Outras coisas, esquece.
        n = int_or(v)
        return 'M' if n == 1 else ('F' if n == 2 else None)

    def code_to_estilo(v):
        n = int_or(v)
        return 'recurvo' if n == 1 else ('composto' if n == 2 else None)

    atletas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or row[0] is None:
            continue  # cabeçalho ou linha vazia

        # Primeira coluna tem de ser ID numérico - se não for, é lixo.
        try:
            atleta_id = str(int(row[0]))
        except (TypeError, ValueError):
            continue

        peso = num(row[1]) if len(row) > 1 else None

        # Altura: aceita metros (1.80) ou centímetros (180). Se vem
        # acima de 3 assumo que está em cm e converto.
        altura_raw = num(row[2]) if len(row) > 2 else None
        if altura_raw is None:
            altura_m = None
        elif altura_raw > 3.0:
            altura_m = altura_raw / 100.0
        else:
            altura_m = altura_raw

        idade = int_or(row[3]) if len(row) > 3 else None
        atleta_estilo = code_to_estilo(row[4]) if len(row) > 4 else None

        # Categoria pode vir como número ou string. Normalizo para
        # lowercase string (juvenil, sénior, etc.) ou None.
        cat_raw = row[5] if len(row) > 5 else None
        if cat_raw is None:
            categoria = None
        elif isinstance(cat_raw, (int, float)):
            categoria = str(int(cat_raw))
        else:
            categoria = (str(cat_raw).strip().lower() or None)

        atleta_genero = code_to_genero(row[6]) if len(row) > 6 else None

        # P1..P30 (scores) e d1..d30 (distâncias) ficam em listas de 30.
        # Onde a célula está vazia, fica None - não tento adivinhar.
        P = [num(row[ci]) if ci < len(row) else None for ci in range(7, 37)]
        d = [num(row[ci]) if ci < len(row) else None for ci in range(38, 68)]

        # P_TOTAL: usa o que está no ficheiro; se faltar, soma P1..P30.
        ptot = num(row[37]) if len(row) > 37 else None
        if ptot is None:
            validos = [p for p in P if p is not None]
            ptot = sum(validos) if validos else None

        atletas.append({
            'id':         atleta_id,
            'nome':       '',   # nome vem da pasta, não está aqui
            'peso_kg':    peso,
            'altura_m':   altura_m,
            'idade':      idade,
            'estilo':     atleta_estilo,
            'categoria':  categoria,
            'genero':     atleta_genero,
            'P':          P,
            'P_total':    ptot,
            'd':          d,
        })

    wb.close()
    return atletas


def atletas_ref_por_id(lista_ref: List[Dict]) -> Dict[str, Dict]:
    """Dict indexado por str(id) a partir da lista devolvida por carregar_atletas_ref."""
    return {str(a['id']): a for a in lista_ref if a.get('id')}


def carregar_confirmacao_arco(caminho):
    """Lê o Inicio_fim_vfinal.xlsx e devolve, por atleta e por ensaio,
    os 3 tempos: toque, confirmação_1, confirmação_2.

    Cada uma das três é uma folha separada. Linhas = atletas, colunas = ensaios.
    """
    from openpyxl import load_workbook
    wb = load_workbook(caminho, read_only=True, data_only=True)

    # As folhas costumam vir com encoding partido (ex: 'confirma��o_1') quando
    # alguém abre o xlsx em Excel num PC com locale errado e salva por cima.
    # Por isso normalizo o nome (lowercase, sem acentos, sem replacement chars)
    # e procuro por palavras-chave + por '1'/'2' isolados.
    folhas = {}  # 'toque' / 'conf_1' / 'conf_2' -> nome real
    for nome in wb.sheetnames:
        n = _normalizar_nome_sheet(nome)
        if 'toque' in n:
            folhas['toque'] = nome
        elif 'confirma' in n:
            if '2' in n:
                folhas['conf_2'] = nome
            elif '1' in n:
                folhas['conf_1'] = nome
            else:
                folhas.setdefault('conf_1', nome)

    def ler_folha(nome_folha):
        """Ler uma folha e devolver {id_atleta: {n_ensaio: tempo_ms}}."""
        ws = wb[nome_folha]
        out = {}
        col_para_ensaio = {}  # coluna -> número do ensaio (1..30)

        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            # Linha 1 é cabeçalho: col 0 vazia, col 1..30 = trials 1..30
            if ri == 1:
                for ci, v in enumerate(row):
                    if ci == 0 or v is None:
                        continue
                    try:
                        col_para_ensaio[ci] = int(v)
                    except (TypeError, ValueError):
                        pass
                continue

            if not row or row[0] is None:
                continue

            try:
                aid = str(int(row[0]))
            except (TypeError, ValueError):
                continue  # ID não numérico, ignorar (provavelmente separador)

            tempos = {}
            for ci, n_ens in col_para_ensaio.items():
                if ci >= len(row) or row[ci] is None:
                    continue
                try:
                    tempos[n_ens] = float(row[ci])
                except (TypeError, ValueError):
                    pass
            if tempos:
                out[aid] = tempos
        return out

    sheets_map = folhas  # alias para o código antigo abaixo
    _ler_aba = ler_folha

    dados_toque  = _ler_aba(sheets_map['toque'])  if 'toque'  in sheets_map else {}
    dados_conf_1 = _ler_aba(sheets_map['conf_1']) if 'conf_1' in sheets_map else {}
    dados_conf_2 = _ler_aba(sheets_map['conf_2']) if 'conf_2' in sheets_map else {}

    # Merge
    por_id: Dict[str, Dict[int, Dict[str, float]]] = {}
    all_ids = set(dados_toque) | set(dados_conf_1) | set(dados_conf_2)
    n_max = 0
    for sid in all_ids:
        por_id[sid] = {}
        trials = set(dados_toque.get(sid, {})) | set(dados_conf_1.get(sid, {})) | set(dados_conf_2.get(sid, {}))
        for t in trials:
            por_id[sid][t] = {
                'toque':  dados_toque.get(sid,  {}).get(t),
                'conf_1': dados_conf_1.get(sid, {}).get(t),
                'conf_2': dados_conf_2.get(sid, {}).get(t),
            }
            n_max = max(n_max, t)

    wb.close()
    return {'por_id': por_id, 'n_trials_max': n_max}


# ═════════════════════════════════════════════════════════════════════
# Analises demograficas (Fase 3 do plano v1.0)
# ═════════════════════════════════════════════════════════════════════

def _valores_chave(atletas, chave):
    """Extrai (atleta, valor_agregado) por atleta - mediana se vier dos ensaios."""
    out = []
    for a in atletas:
        v = a.get(chave) if isinstance(a, dict) else None
        if v is None:
            vs = []
            mets = a.get('mets') if isinstance(a, dict) else None
            if isinstance(mets, dict):
                for _lado, lista in mets.items():
                    if isinstance(lista, list):
                        for m in lista:
                            if isinstance(m, dict) and m.get(chave) is not None:
                                try:
                                    vs.append(float(m[chave]))
                                except (TypeError, ValueError):
                                    pass
            if vs:
                try:
                    v = float(np.median(vs))
                except Exception:
                    v = None
        if v is None:
            continue
        try:
            fv = float(v)
            if math.isnan(fv) or math.isinf(fv):
                continue
            out.append((a, fv))
        except (TypeError, ValueError):
            continue
    return out


def comparar_grupos(atletas, chave, fator):
    """Mann-Whitney (2 grupos) ou Kruskal-Wallis (3+ grupos)."""
    pares = _valores_chave(atletas, chave)
    grupos: Dict[Any, List[float]] = {}
    for a, v in pares:
        g = a.get(fator)
        if g is None or g == '':
            continue
        if isinstance(g, str):
            g = g.strip()
            if fator == 'genero':
                g = g.upper()
            else:
                g = g.lower()
        grupos.setdefault(g, []).append(v)
    grupos = {k: vs for k, vs in grupos.items() if len(vs) >= 2}
    n_por_grupo = {k: len(vs) for k, vs in grupos.items()}
    mediana = {k: float(np.median(vs)) for k, vs in grupos.items()}
    p25 = {k: float(np.percentile(vs, 25)) for k, vs in grupos.items()}
    p75 = {k: float(np.percentile(vs, 75)) for k, vs in grupos.items()}
    teste = estat = pval = None
    n_grupos = len(grupos)
    try:
        if n_grupos == 2:
            chaves = list(grupos.keys())
            U, p = _sp_stats.mannwhitneyu(grupos[chaves[0]], grupos[chaves[1]],
                                           alternative='two-sided')
            teste, estat, pval = 'mannwhitneyu', float(U), float(p)
        elif n_grupos >= 3:
            H, p = _sp_stats.kruskal(*[grupos[k] for k in grupos])
            teste, estat, pval = 'kruskal', float(H), float(p)
    except Exception:
        pass
    return {
        'fator': fator, 'chave': chave, 'grupos': grupos,
        'n_por_grupo': n_por_grupo, 'mediana': mediana,
        'p25': p25, 'p75': p75,
        'teste': teste, 'estatistica': estat, 'p_valor': pval,
        'n_grupos': n_grupos,
    }


def correlacao_demografica(atletas, chave_cop, chave_dem):
    """Pearson + Spearman entre metrica CoP e variavel demografica continua."""
    pares_cop = _valores_chave(atletas, chave_cop)
    xs, ys = [], []
    for a, cop_v in pares_cop:
        dem_v = a.get(chave_dem) if isinstance(a, dict) else None
        if dem_v is None:
            continue
        try:
            dem_f = float(dem_v)
            if math.isnan(dem_f) or math.isinf(dem_f):
                continue
        except (TypeError, ValueError):
            continue
        xs.append(dem_f); ys.append(cop_v)
    pr = pp = sr = sp = None
    if len(xs) >= 3:
        try:
            r, p = _sp_stats.pearsonr(xs, ys)
            pr, pp = float(r), float(p)
        except Exception:
            pass
        try:
            r, p = _sp_stats.spearmanr(xs, ys)
            sr, sp = float(r), float(p)
        except Exception:
            pass
    return {
        'chave_cop': chave_cop, 'chave_dem': chave_dem,
        'n': len(xs),
        'pearson_r': pr, 'pearson_p': pp,
        'spearman_r': sr, 'spearman_p': sp,
        'x': xs, 'y': ys,
    }


def percentis_subgrupo(ath, atletas, chave, fatores=('categoria', 'genero')):
    """Posicao do atleta dentro do subgrupo demografico (P25/P50/P75/percentil/rank)."""
    if not isinstance(ath, dict):
        return {'valor_atleta': None, 'subgrupo': {}, 'n_subgrupo': 0,
                'p25': None, 'p50': None, 'p75': None,
                'percentil_atleta': None, 'rank': None}
    subgrupo = {f: ath.get(f) for f in fatores}

    def _bate(a):
        for f in fatores:
            va = ath.get(f); vb = a.get(f)
            if va is None and vb is None:
                continue
            if va is None or vb is None:
                return False
            if isinstance(va, str) and isinstance(vb, str):
                if va.strip().lower() != vb.strip().lower():
                    return False
            elif va != vb:
                return False
        return True

    pares_sub = [p for p in _valores_chave(atletas, chave) if _bate(p[0])]
    valores = [v for _, v in pares_sub]
    n_sub = len(valores)

    v_ath = None
    for _, v in _valores_chave([ath], chave):
        v_ath = v; break

    p25 = p50 = p75 = None
    pct = rank = None
    if n_sub >= 3:
        p25 = float(np.percentile(valores, 25))
        p50 = float(np.percentile(valores, 50))
        p75 = float(np.percentile(valores, 75))
        if v_ath is not None:
            n_le = sum(1 for v in valores if v <= v_ath)
            pct = round(100.0 * n_le / n_sub, 1)
            sorted_vals = sorted(valores)
            rank = 1
            for v in sorted_vals:
                if v < v_ath:
                    rank += 1
                else:
                    break

    return {
        'valor_atleta': v_ath, 'subgrupo': subgrupo, 'n_subgrupo': n_sub,
        'p25': p25, 'p50': p50, 'p75': p75,
        'percentil_atleta': pct, 'rank': rank,
    }


def correlacao_score(atletas, chave_cop, chave_outcome='P_total'):
    """
    Correlaciona metrica CoP com score. Devolve:
      {'agregado': {...}, 'per_ensaio': {...} ou None}
    """
    agreg = correlacao_demografica(atletas, chave_cop, chave_outcome)
    xs, ys = [], []
    for a in atletas:
        P_list = a.get('P') if isinstance(a, dict) else None
        if not isinstance(P_list, list):
            continue
        mets = a.get('mets') if isinstance(a, dict) else None
        if not isinstance(mets, dict):
            continue
        ensaios = mets.get('arco') or []
        for i, m in enumerate(ensaios):
            if not isinstance(m, dict):
                continue
            if i >= len(P_list):
                continue
            p_val = P_list[i]
            cop_val = m.get(chave_cop)
            if p_val is None or cop_val is None:
                continue
            try:
                xs.append(float(p_val)); ys.append(float(cop_val))
            except (TypeError, ValueError):
                continue
    per = None
    if len(xs) >= 3:
        try:
            pr, pp = _sp_stats.pearsonr(xs, ys)
            sr, sp = _sp_stats.spearmanr(xs, ys)
            per = {
                'n': len(xs),
                'pearson_r': float(pr), 'pearson_p': float(pp),
                'spearman_r': float(sr), 'spearman_p': float(sp),
            }
        except Exception:
            per = {'n': len(xs), 'pearson_r': None, 'pearson_p': None,
                   'spearman_r': None, 'spearman_p': None}
    return {
        'chave_cop': chave_cop, 'chave_outcome': chave_outcome,
        'agregado': agreg, 'per_ensaio': per,
    }


# ═════════════════════════════════════════════════════════════════════
# Testes
# ═════════════════════════════════════════════════════════════════════

def run_testes(verbose: bool = True) -> bool:
    """
    Testes sinteticos para o modulo bsp_core. Devolve True se todos passarem.

    Cobre:
      - calcular() com janela e sem janela
      - Normalizacoes altura/massa
      - Outliers e assimetria
      - Demografia (comparar_grupos, correlacao_demografica, percentis, score)
    """
    resultados: List[Tuple[str, bool]] = []

    def _t(nome, ok, detalhe=''):
        resultados.append((nome, bool(ok)))
        if verbose:
            sym = '[ok]' if ok else '[FAIL]'
            print(f'  {sym} {nome}' + (f'  ({detalhe})' if detalhe else ''))

    if verbose:
        print(f"\n{'='*58}\n  bsp_core v{VERSAO} - Testes\n{'='*58}\n")

    # ── [1] Constantes e imports ──────────────────────────────────────
    if verbose: print('  [1] Constantes e imports')
    _t('PROTO_FMS definido', PROTO_FMS == 'fms')
    _t('PROTO_ARCO definido', PROTO_ARCO == 'tiro_arco')
    _t('G_ACEL ~ 9.80665', abs(G_ACEL - 9.80665) < 1e-6, f'={G_ACEL}')
    _t('PROTOCOLOS[PROTO_ARCO] n_ens=30',
       PROTOCOLOS[PROTO_ARCO]['n_ens'] == 30)
    # Nao deve importar tkinter/matplotlib/reportlab
    import sys as _sys
    mods = set(_sys.modules)
    banned = {m for m in mods if m.startswith('tkinter')
              or m.startswith('matplotlib') or m.startswith('reportlab')
              or m == 'PIL'}
    _t('import nao puxa tkinter/matplotlib/reportlab/PIL',
       len(banned) == 0, f'banned={banned}' if banned else '')

    # ── [2] calcular() ─────────────────────────────────────────────────
    if verbose: print('\n  [2] calcular()')
    # Frames sinteticos: 100 samples a 100 Hz
    frames = []
    for i in range(100):
        t_ms = i * 10.0
        frames.append({
            't_ms': t_ms,
            'x': 10.0 * math.sin(2*math.pi*1.0*t_ms/1000.0),
            'y': 15.0 * math.sin(2*math.pi*0.5*t_ms/1000.0),
        })
    m = calcular(frames)
    _t('calcular devolve dict', isinstance(m, dict))
    _t('amp_x > 0', m['amp_x'] > 0, f"amp_x={m['amp_x']:.2f}")
    _t('amp_y > 0', m['amp_y'] > 0, f"amp_y={m['amp_y']:.2f}")
    _t('ea95 > 0',  m['ea95']  > 0, f"ea95={m['ea95']:.2f}")
    _t('rms_r > 0', m['rms_r'] > 0)
    _t('len(cof_x) == 100', len(m['cof_x']) == 100)
    # Com peso e altura (parametros mantidos por compatibilidade)
    m2 = calcular(frames, peso_kg=70.0, altura_m=1.75)

    # Janela
    m3 = calcular(frames, t_ini=200, t_fim=800)
    _t('janela [200,800] reduz frames', len(m3['cof_x']) < 100,
       f"n={len(m3['cof_x'])}")

    # Muito poucos frames
    m4 = calcular([{'t_ms': 0, 'x': 1, 'y': 1},
                   {'t_ms': 10, 'x': 2, 'y': 2}])
    _t('< 5 frames devolve None', m4 is None)

    # ── [3] assimetria e outliers ─────────────────────────────────────
    if verbose: print('\n  [3] assimetria e outliers')
    _t('assimetria(100, 80) > 0', assimetria(100, 80) > 0,
       f'ai={assimetria(100, 80)}')
    _t('assimetria(None, X) = None', assimetria(None, 5) is None)
    _t('assimetria(5, 5) = 0', assimetria(5, 5) == 0.0)

    mets = [{'ea95': 100}, {'ea95': 105}, {'ea95': 98},
            {'ea95': 103}, {'ea95': 1000}]  # ultimo e outlier
    flags = flagrar_outliers(mets, chave='ea95', z_thresh=3.0, pct_thresh=0.3)
    _t('flagrar_outliers deteca outlier obvio', flags[-1] is True,
       f'flags={flags}')
    _t('flagrar_outliers nao flagra normais',
       all(not f for f in flags[:-1]))

    # ── [4] Demografia ─────────────────────────────────────────────────
    if verbose: print('\n  [4] Demografia')
    # 12 atletas: 6M (ea95 ~400) + 6F (ea95 ~200)
    atletas_demo = []
    for i in range(6):
        atletas_demo.append({
            'id': f'M{i+1}', 'nome': f'Atleta M{i+1}',
            'genero': 'M', 'categoria': 'senior', 'estilo': 'recurvo',
            'peso_kg': 75.0 + i, 'altura_m': 1.80, 'idade': 30,
            'ea95': 400.0 + i*10,
            'P': [8+i//2]*30, 'P_total': (8+i//2)*30,
            'mets': {'arco': [{'ea95': 400.0 + i*10} for _ in range(30)]},
        })
    for i in range(6):
        atletas_demo.append({
            'id': f'F{i+1}', 'nome': f'Atleta F{i+1}',
            'genero': 'F', 'categoria': 'senior', 'estilo': 'composto',
            'peso_kg': 60.0 + i, 'altura_m': 1.65, 'idade': 28,
            'ea95': 200.0 + i*10,
            'P': [9+i//2]*30, 'P_total': (9+i//2)*30,
            'mets': {'arco': [{'ea95': 200.0 + i*10} for _ in range(30)]},
        })

    r_gen = comparar_grupos(atletas_demo, 'ea95', 'genero')
    _t('comparar_grupos(genero): 2 grupos',
       r_gen['n_grupos'] == 2, f"n={r_gen['n_grupos']}")
    _t('comparar_grupos(genero): teste=mannwhitneyu',
       r_gen['teste'] == 'mannwhitneyu')
    _t('comparar_grupos(genero): p<0.05',
       r_gen['p_valor'] is not None and r_gen['p_valor'] < 0.05,
       f"p={r_gen['p_valor']:.4g}" if r_gen['p_valor'] else '')

    r_cor = correlacao_demografica(atletas_demo, 'ea95', 'peso_kg')
    _t('correlacao_demografica: n=12', r_cor['n'] == 12)
    _t('correlacao_demografica: pearson_r nao None',
       r_cor['pearson_r'] is not None)
    _t('correlacao_demografica: r positivo',
       r_cor['pearson_r'] > 0, f"r={r_cor['pearson_r']:.3f}")

    pct = percentis_subgrupo(atletas_demo[2], atletas_demo, 'ea95')
    _t('percentis_subgrupo: n_subgrupo=6 (6M seniores)',
       pct['n_subgrupo'] == 6)
    _t('percentis_subgrupo: rank entre 1..6',
       pct['rank'] is not None and 1 <= pct['rank'] <= 6)

    cs = correlacao_score(atletas_demo, 'ea95', 'P_total')
    _t('correlacao_score: agregado.n=12',
       cs['agregado']['n'] == 12)
    _t('correlacao_score: per_ensaio n>=360',
       cs['per_ensaio'] is not None and cs['per_ensaio']['n'] >= 360)

    # ── [5] Indice de oscilacao ───────────────────────────────────────
    if verbose: print('\n  [5] Indice de oscilacao')
    _t('stiff_x calculado', m['stiff_x'] is not None, f"stiff_x={m.get('stiff_x')}")

    # ── [6] Parser arco ────────────────────────────────────────────────
    if verbose: print('\n  [6] Parser Tiro com Arco')
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        p = os.path.join(td, '101_1 - 03-06-2025 - Stability export.xls')
        # Cria ficheiro tab-separated com cabecalho minimo
        with open(p, 'w', encoding='iso-8859-1') as f:
            f.write('Stability export for measurement: test\n')
            f.write('Patient name: Teste\n')
            f.write('Measurement done on 03-06-2025\n')
            f.write('\n')
            f.write('Frame\tTime (ms)\tEntire plate COF X\tEntire plate COF Y\tEntire plate COF Force\n')
            for i in range(100):
                f.write(f'{i}\t{i*20.0}\t{10.0*math.sin(i*0.1)}\t{5.0*math.cos(i*0.1)}\t500.0\n')

        _t('_detectar_formato_arco True', _detectar_formato_arco(p))
        info = _ler_ficheiro_arco(p)
        _t('ler_arco devolve frames', len(info['frames']) == 100)
        _t('ler_arco paciente detectado',
           info['paciente'] == 'Teste')
        _t('ler_arco data detectada',
           info['data'] == '03-06-2025')

        # Matcher
        achado = achar_ficheiro_arco(td, '101', 1)
        _t('achar_ficheiro_arco encontra ficheiro',
           achado is not None and achado.endswith('Stability export.xls'))

        # Regressao: cabecalho com sufixo de unidade '(mm)', '(N)' ⇒ deve ler
        # (formato real do ficheiro Stability export.xls do platform export)
        p2 = os.path.join(td, '202_1 - 06-11-2023 - Stability export.xls')
        with open(p2, 'w', encoding='iso-8859-1') as f:
            f.write('Stability export for measurement: 202_1\n')
            f.write('Patient name: Teste 2\n')
            f.write('Measurement done on 06-11-2023\n')
            f.write('\n\n')
            f.write('Interval\tBegin (ms)\tEnd (ms)\n')
            f.write('1\t0\t1000\n\n')
            f.write('Frame\tTime (ms)\tEntire plate COF X (mm)\tEntire plate COF Y (mm)\tEntire plate Force (N)\tEntire plate COF Frame distance (mm)\n')
            for i in range(50):
                f.write(f'{i}\t{i*20.0}\t{628.0+0.1*i}\t{136.0-0.05*i}\t{1060+i}\t{0.5}\n')
        info2 = _ler_ficheiro_arco(p2)
        _t('ler_arco aceita cabecalho "(mm)" / "(N)" (regressao Stability export real)',
           len(info2['frames']) == 50 and info2['frames'][0]['x'] == 628.0,
           f"n={len(info2['frames'])}, x0={info2['frames'][0].get('x') if info2['frames'] else None}")

        # Regressao: ficheiro 'Entire plate roll off.xls' nao deve ser apanhado
        roll_off = os.path.join(td, '202_1 - 06-11-2023 - Entire plate roll off.xls')
        open(roll_off, 'w').write('dummy')  # ficheiro qualquer
        achado2 = achar_ficheiro_arco(td, '202', 1)
        _t('achar_ficheiro_arco IGNORA roll-off, devolve Stability export',
           achado2 is not None and 'roll off' not in achado2.lower(),
           f"achou={os.path.basename(achado2) if achado2 else None}")

        # Regressao: carregar_atletas_ref com xlsx no formato real
        # (col 1=ID, col 2=PESO, col 3=ALTURA, ..., sem coluna Nome)
        try:
            from openpyxl import Workbook as _Wb
            xref = os.path.join(td, 'ref.xlsx')
            wb_r = _Wb()
            ws_r = wb_r.active
            ws_r.append(['', 'PESO', 'ALTURA', 'IDADE', 'ESTILO', 'CATEGORIA', 'GENERO']
                        + [f'P{i}' for i in range(1, 31)] + ['P_TOTAL']
                        + [f'd{i}' for i in range(1, 31)] + ['', 'duracao_1'])
            ws_r.append([101, 84.5, 1.8, 21, 1, 3, 1] + [9]*30 + [270] + [25.0]*30 + [101, 5.02])
            ws_r.append([102, 60.5, 1.62, 18, 2, 1, 2] + [8]*30 + [240] + [30.0]*30 + [102, 4.50])
            wb_r.save(xref)

            refs_t = carregar_atletas_ref(xref)
            _t('carregar_atletas_ref: 2 atletas (regressao formato sem coluna Nome)',
               len(refs_t) == 2, f"n={len(refs_t)}")
            r1 = refs_t[0]
            _t('carregar_atletas_ref: atleta 101 peso 84.5 kg, altura 1.8 m',
               r1['peso_kg'] == 84.5 and r1['altura_m'] == 1.8,
               f"peso={r1['peso_kg']} altura={r1['altura_m']}")
            _t('carregar_atletas_ref: estilo recurvo, genero M',
               r1['estilo'] == 'recurvo' and r1['genero'] == 'M',
               f"estilo={r1['estilo']} genero={r1['genero']}")
            _t('carregar_atletas_ref: P_total e os 30 P em ordem',
               r1['P_total'] == 270 and len(r1['P']) == 30 and r1['P'][0] == 9,
               f"P_total={r1['P_total']}, n_P={len(r1['P'])}, P[0]={r1['P'][0]}")
            _t('carregar_atletas_ref: 30 distancias',
               len(r1['d']) == 30 and r1['d'][0] == 25.0,
               f"n_d={len(r1['d'])}, d[0]={r1['d'][0]}")
            r2 = refs_t[1]
            _t('carregar_atletas_ref: atleta 102 estilo composto, genero F',
               r2['estilo'] == 'composto' and r2['genero'] == 'F',
               f"estilo={r2['estilo']} genero={r2['genero']}")
        except Exception as ex_ref:
            _t('carregar_atletas_ref regressao sem excepcao',
               False, f"{type(ex_ref).__name__}: {str(ex_ref)[:80]}")

        # Regressao: carregar_confirmacao_arco com 3 folhas e nomes contendo
        # 'confirmação_1' / 'confirmação_2' (que tem 'i' em comum -> bug fixed)
        try:
            xj = os.path.join(td, 'janelas.xlsx')
            wb_j = _Wb()
            wb_j.remove(wb_j.active)
            ws_t = wb_j.create_sheet('tempo do toque')
            ws_t.append([None] + list(range(1, 6)))   # cabecalho: trials 1..5
            ws_t.append([101, 300, 180, 500, 260, 0])
            ws_t.append([102, 560, 500, 860, 480, 1480])
            ws_c1 = wb_j.create_sheet('confirmação_1')
            ws_c1.append([None] + list(range(1, 6)))
            ws_c1.append([101, 1500, 1460, 1180, 860, 1000])
            ws_c1.append([102, 3240, 4180, 4020, 2960, 3640])
            ws_c2 = wb_j.create_sheet('confirmação_2')
            ws_c2.append([None] + list(range(1, 6)))
            ws_c2.append([101, 6520, 5040, 4120, 4760, 4500])
            ws_c2.append([102, 4900, 5800, 7440, 4940, 6060])
            wb_j.save(xj)

            jres = carregar_confirmacao_arco(xj)
            _t('carregar_confirmacao_arco: 2 atletas',
               len(jres['por_id']) == 2, f"n={len(jres['por_id'])}")
            _t('carregar_confirmacao_arco: max_trials=5 (regressao eixo)',
               jres['n_trials_max'] == 5, f"got={jres['n_trials_max']}")
            j101_1 = jres['por_id'].get('101', {}).get(1, {})
            _t('carregar_confirmacao_arco: 101 trial 1 toque=300',
               j101_1.get('toque') == 300, f"got={j101_1.get('toque')}")
            _t('carregar_confirmacao_arco: 101 trial 1 conf_1=1500',
               j101_1.get('conf_1') == 1500, f"got={j101_1.get('conf_1')}")
            _t('carregar_confirmacao_arco: 101 trial 1 conf_2=6520 (regressao "i" match)',
               j101_1.get('conf_2') == 6520, f"got={j101_1.get('conf_2')}")
        except Exception as ex_j:
            _t('carregar_confirmacao_arco regressao sem excepcao',
               False, f"{type(ex_j).__name__}: {str(ex_j)[:80]}")

    # ── Relatorio ──────────────────────────────────────────────────────
    total = len(resultados)
    n_ok  = sum(1 for _, ok in resultados if ok)
    n_fail = total - n_ok
    if verbose:
        print(f"\n{'='*58}")
        print(f"  RESULTADO: {n_ok}/{total} testes passaram")
        if n_fail > 0:
            print(f"  FALHAS ({n_fail}):")
            for n, ok in resultados:
                if not ok:
                    print(f"    [FAIL] {n}")
        else:
            print('  Todos os testes passaram!')
        print(f"{'='*58}")
    return n_fail == 0


if __name__ == '__main__':
    import argparse, sys
    p = argparse.ArgumentParser(description='bsp_core - modulo nucleo do BSP')
    p.add_argument('--testes', action='store_true',
                   help='Correr suite de testes sinteticos')
    args = p.parse_args()
    if args.testes:
        ok = run_testes(verbose=True)
        sys.exit(0 if ok else 1)
    print(f'bsp_core v{VERSAO} - ver {__file__}')
    print('Use --testes para correr os testes.')
